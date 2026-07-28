import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


class ExcelGenerator:
    """
    Generates Excel (.xlsx) reports.
    """

    def _style_header(self, worksheet):
        """
        Apply styling to header row.
        """

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

    def generate_detection_report(
        self,
        detections,
        output_path: str,
    ):
        """
        Generate Detection Excel Report.
        """

        os.makedirs(
            os.path.dirname(output_path),
            exist_ok=True,
        )

        workbook = Workbook()

        sheet = workbook.active

        if sheet is None:
            raise RuntimeError("Failed to create worksheet")

        sheet.title = "Detections"

        sheet.append(
            [
                "ID",
                "Camera",
                "Object",
                "Confidence",
                "X1",
                "Y1",
                "X2",
                "Y2",
                "Detected At",
            ]
        )

        self._style_header(sheet)

        for detection in detections:
            sheet.append(
                [
                    detection.id,
                    detection.camera_id,
                    detection.label,
                    detection.confidence,
                    detection.x1,
                    detection.y1,
                    detection.x2,
                    detection.y2,
                    str(detection.detected_at),
                ]
            )

        workbook.save(output_path)

        return output_path

    def generate_alert_report(
        self,
        alerts,
        output_path: str,
    ):
        """
        Generate Alert Excel Report.
        """

        os.makedirs(
            os.path.dirname(output_path),
            exist_ok=True,
        )

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Alerts"

        sheet.append(
            [
                "ID",
                "Camera",
                "Detection",
                "Title",
                "Severity",
                "Status",
                "Created At",
            ]
        )

        self._style_header(sheet)

        for alert in alerts:
            sheet.append(
                [
                    alert.id,
                    alert.camera_id,
                    alert.detection_id,
                    alert.title,
                    alert.severity,
                    alert.status,
                    str(alert.created_at),
                ]
            )

        workbook.save(output_path)

        return output_path

    def generate_recording_report(
        self,
        recordings,
        output_path: str,
    ):
        """
        Generate Recording Excel Report.
        """

        os.makedirs(
            os.path.dirname(output_path),
            exist_ok=True,
        )

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Recordings"

        sheet.append(
            [
                "ID",
                "Camera",
                "File Name",
                "Duration",
                "Size (MB)",
                "Start Time",
                "End Time",
                "Created At",
            ]
        )

        self._style_header(sheet)

        for recording in recordings:
            sheet.append(
                [
                    recording.id,
                    recording.camera_id,
                    recording.file_name,
                    recording.duration,
                    recording.size,
                    str(recording.start_time),
                    str(recording.end_time),
                    str(recording.created_at),
                ]
            )

        workbook.save(output_path)

        return output_path
